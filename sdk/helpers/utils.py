import json

import requests
from aiohttp import ClientSession
from aiohttp_socks import ProxyConnector
from openpyxl import Workbook
from starknet_py.net.full_node_client import FullNodeClient
from starknet_py.net.gateway_client import GatewayClient
from starknet_py.net.networks import MAINNET

from config import (
    STARKNET_RPC_PROVIDER,
    USE_PROXY,
    STARKNET_ETH_MIN_BALANCE,
    IP_CHANGE_LINK,
    USE_MOBILE_PROXY,
    ATTEMPTS_COUNT
)
from constants import (
    STARKNET_TOKENS,
    STARKNET_ETH_TOKEN_ADDRESS,
    COINGECKO_TOKEN_IDS
)
from sdk.helpers.decorators import starknet_retry
from sdk.helpers.logger import logger
from sdk.starknet.client import StarknetClient
from sdk.starknet.utils import get_cg_tokens_price_usd, wei_to_float


def export_json(data: str, field: str, destination):
    try:
        json_data = json.loads(data)

        wb = Workbook()
        ws = wb.active

        if len(json_data[field]) == 0:
            return None

        headers = list(json_data[field][0].keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(json_data["data"], start=2):
            for col_idx, key in enumerate(headers, start=1):
                value = row_data[key]
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(destination)

    except Exception as e:
        logger.error(f"Encountered an error while exporting db to Excel: {str(e)}")
        exit()


async def close_starknet_session(starknet_read_client=None, starknet_write_client=None):
    if USE_PROXY:
        if starknet_read_client and starknet_read_client.rpc._client.session is not None:
            await starknet_read_client.rpc._client.session.close()
            await starknet_read_client.rpc._client.session.close()

        if starknet_write_client and starknet_write_client.rpc._feeder_gateway_client.session is not None:
            await starknet_write_client.rpc._feeder_gateway_client.session.close()
            await starknet_write_client.rpc._gateway_client.session.close()


def read_from_json(file_path):
    try:
        with open(file_path) as json_file:
            return json.load(json_file)

    except FileNotFoundError:
        logger.error(f"File '{file_path}' not found")
        exit()

    except Exception as e:
        logger.error(f"Encountered an error while reading a JSON file '{file_path}': {e}")
        exit()


def change_mobile_ip() -> None:
    try:
        if USE_MOBILE_PROXY:
            res = requests.get(IP_CHANGE_LINK)

            if res.status_code == 200:
                logger.info("IP address changed successfully", send_to_tg=False)
            else:
                raise Exception("Failed to change IP address")

    except Exception as e:
        raise Exception(f"Encountered an error when changing ip address, check your proxy provider: {e}")


def get_starknet_read_client(proxy):
    if USE_PROXY:
        return FullNodeClient(
            node_url=STARKNET_RPC_PROVIDER,
            session=ClientSession(connector=ProxyConnector.from_url(f'http://{proxy}'))
        )
    else:
        return FullNodeClient(node_url=STARKNET_RPC_PROVIDER)


def get_starknet_write_client(proxy):
    if not USE_PROXY:
        return GatewayClient(net=MAINNET)

    return GatewayClient(
        net=MAINNET,
        session=ClientSession(connector=ProxyConnector.from_url(f'http://{proxy}'))
    )


async def get_starknet_max_balance_token(client: StarknetClient) -> (str, int):
    usd_token_prices = get_cg_tokens_price_usd(COINGECKO_TOKEN_IDS)
    max_balance_token_address = None
    max_balance_token_amount = 0.0
    max_balance_token_usd = 0.0

    for token_address in STARKNET_TOKENS:
        token_balance = await get_starknet_token_balance(client, token_addr=token_address)
        token_balance_usd = token_balance * usd_token_prices[STARKNET_TOKENS.index(token_address)]

        if max_balance_token_usd < token_balance_usd:
            max_balance_token_amount = token_balance
            max_balance_token_usd = token_balance_usd
            max_balance_token_address = token_address

    if max_balance_token_address == STARKNET_ETH_TOKEN_ADDRESS:
        max_balance_token_amount -= STARKNET_ETH_MIN_BALANCE

    if max_balance_token_address is None:
        raise Exception("Max balance token address is None")

    if max_balance_token_amount is None:
        raise Exception("Max balance token amount is None")

    if max_balance_token_amount == 0.0:
        raise Exception("Max balance token amount is 0")

    return max_balance_token_address, max_balance_token_amount


@starknet_retry(attempts=ATTEMPTS_COUNT)
async def get_starknet_token_balance(client: StarknetClient, token_addr=STARKNET_ETH_TOKEN_ADDRESS) -> float:
    try:
        balance = await client.get_balance(token_address=token_addr)
        return wei_to_float(balance, token_addr)
    except Exception as e:
        raise Exception(f"Client failed with error: {str(e)}")


def greeting_message():
    bridge_usage_warning_message = ("Если вы гоняете пачку аккаунтов, то ОБЯЗАТЕЛЬНО нужно протестить сначала на ОДНОМ "
                                    "кошельке с МАЛЕНЬКОЙ суммой, удостовериться, что деньги дошли, и ТОЛЬКО ПОСЛЕ "
                                    "ЭТОГО делать на пачку, просматривая, в идеале, что все дошло.")
    start_message = fr'''

                                  ^Y                  
                                 ~&@7                
                      75~:.     !@&~&:       , .      
                      .&&PYY7^.7@@# J#   .^7JPB^      
                       ^@&Y:^?Y&@@P  GBB&@@GP&~       
                        7@@&?  :&@J  G@@&Y.~#^        
                     .:~?&#&@&? !@! B@G~  !&:         
                :75PPY?!^. .:?GG~P!5P~^!YG@@GJ~.      
                .~YG#&&##B#BGPJ?J??J?J5GBBBB##&#B5!.  
                    .^?P&@BJ!^^5G~G^5GJ^. .:!?Y5P57:\  
                       .#?  ^P@#.:@J !&@@#&J~^.       
                      :#7.J&@@#. !@@~  !&@@5          
                     ^&GP@@&BG#. J@@@5?~:?&@7         
                    :BGJ7^..  GG P@@J.:!JY5&@:        
                    .         .&7B@?      .~YJ        
                              ^&@7                   
                                ?!                  

               __    _ __                        __                  
   _______  __/ /_  (_) /  _   __   ____  ____ _/ /______  ____  ___ 
  / ___/ / / / __ \/ / /  | | / /  /_  / / __ `/ //_/ __ \/ __ \/ _ \
 (__  ) /_/ / /_/ / / /   | |/ /    / /_/ /_/ / ,< / /_/ / / / /  __/   
/____/\__, /_.___/_/_/    |___/    /___/\__,_/_/|_|\____/_/ /_/\___/ 
     /____/                                                          

Modules:
 1: Create database
 2: Warmup (all contracts)
 3: Warmup with gas (all contracts)
 4: Warmup low bank (all contracts)
 5: Collector (tokens -> eth)
 6: Eth sender (wallet -> okx)
 7: Eth withdrawal to starknet (okx -> wallet)
 8: Starkgate (eth -> stark). {bridge_usage_warning_message}
 9: Starkgate (stark -> eth). {bridge_usage_warning_message}
10: Orbiter (arb -> stark). {bridge_usage_warning_message}
11: Orbiter (stark -> arb). {bridge_usage_warning_message}
12: Layerswap (op -> stark). {bridge_usage_warning_message}
13: Layerswap (stark -> op). {bridge_usage_warning_message}
14: (TODO) Cairo 1 update (argentx)
    '''

    logger.success(start_message, send_to_tg=False)
